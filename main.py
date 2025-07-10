import csv
import json
import qrcode
import openpyxl
import os
from datetime import datetime
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.properties import StringProperty
from kivy.uix.image import Image
import cv2
from pyzbar.pyzbar import decode
import json


# File paths
CSV_FILE = 'current_gatepass.csv'
EXCEL_LOG = 'gatepass_log.xlsx'
QR_IMAGE = 'gatepass_qr.png'
EMP_EXCEL = 'emp-info.xlsx'

current_user_id = None
current_user_role = None
current_user_name = None
employee_info = {}

def load_credentials_from_excel(path=EMP_EXCEL):
    creds = {}
    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            uid, name, pwd, role, mgr, hr_ids = row
            uid_str = str(uid).strip()
            mgr_str = str(mgr).strip() if mgr else ''
            hr_list = [x.strip() for x in str(hr_ids).split(',')] if hr_ids else []

            creds[uid_str] = {
                'password': str(pwd).strip(),
                'role': role.strip().lower() if role else 'employee',
                'name': name,
                'manager': mgr_str,
                'hr_ids': hr_list
            }
            employee_info[uid_str] = {
                'name': name,
                'role': role.strip().lower() if role else 'employee',
                'manager': mgr_str,
                'hr_ids': hr_list
            }
    except Exception as e:
        print("Excel load error:", e)
    return creds

def load_data():
    try:
        with open(CSV_FILE, 'r') as f:
            return list(csv.DictReader(f))
    except FileNotFoundError:
        return []

def save_data(data):
    try:
        file_exists = os.path.exists(CSV_FILE)
        with open(CSV_FILE, 'a', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=data.keys())
            if not file_exists or os.stat(CSV_FILE).st_size == 0:
                writer.writeheader()
            writer.writerow(data)
    except Exception as e:
        print("Write error:", e)

def save_to_excel(reason):
    filename = EXCEL_LOG
    headers = ["Date", "Time", "Emp-id", "Emp-name", "Reason", "Manager-approval", "HR-approval"]
    row = [
        datetime.now().strftime("%Y-%m-%d"),
        datetime.now().strftime("%H:%M:%S"),
        current_user_id,
        current_user_name,
        reason,
        "",
        ""
    ]
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(headers)
    sheet.append(row)
    wb.save(filename)

def get_manager_chain(emp_id):
    chain = []
    current = emp_id
    while True:
        manager_id = employee_info.get(current, {}).get('manager')
        if manager_id and manager_id not in chain:
            chain.append(manager_id)
            current = manager_id
        else:
            break
    return chain

def get_hr_chain(emp_id):
    chain = []
    hr_ids = employee_info.get(emp_id, {}).get('hr_ids', [])
    for hr_id in hr_ids:
        if hr_id not in chain:
            chain.append(hr_id)
            chain.extend(get_manager_chain(hr_id))
    return list(set(chain))

class LoginScreen(Screen):
    def login(self):
        global current_user_id, current_user_role, current_user_name
        uid = self.ids.user_input.text.strip()
        pwd = self.ids.pwd_input.text.strip()
        creds = load_credentials_from_excel()
        user = creds.get(uid)
        if user and user['password'] == pwd:
            current_user_id = uid
            current_user_role = user['role']
            current_user_name = user['name']
            self.manager.current = 'main'
        else:
            self.ids.message.text = "\u274C Invalid credentials"

class MainScreen(Screen):
    user_role = StringProperty("employee")

    def on_pre_enter(self):
        self.user_role = current_user_role

        # Always show Scan & Exit
        self.ids.btn_scan.opacity = 1
        self.ids.btn_scan.disabled = False
        self.ids.btn_exit.opacity = 1
        self.ids.btn_exit.disabled = False

        if self.user_role == 'security':
            # Hide Gatepass Entry, View Gatepass, Approval
            self.ids.btn_entry.opacity = 0
            self.ids.btn_entry.disabled = True

            self.ids.btn_view.opacity = 0
            self.ids.btn_view.disabled = True

            self.ids.btn_approval.opacity = 0
            self.ids.btn_approval.disabled = True
        else:
            # Show all buttons for non-security roles
            self.ids.btn_entry.opacity = 1
            self.ids.btn_entry.disabled = False

            self.ids.btn_view.opacity = 1
            self.ids.btn_view.disabled = False

            self.ids.btn_approval.opacity = 1
            self.ids.btn_approval.disabled = False

    def welcome_text(self):
        return f"Welcome, {current_user_name or ''}"

    def logout(self):
        global current_user_id, current_user_role, current_user_name
        current_user_id = None
        current_user_role = None
        current_user_name = None
        self.manager.current = 'login'

    def go_to_approval(self):
        if self.user_role == 'manager':
            self.manager.current = 'manager'
        else:
            self.manager.current = 'hr'


class EntryScreen(Screen):
    def on_pre_enter(self):
        self.ids.empid.text = current_user_id or ''
        self.ids.name.text = current_user_name or ''
        self.ids.time.text = datetime.now().strftime("%H:%M:%S")
        self.ids.reason.text = ''

    def submit(self):
        reason = self.ids.reason.text.strip()
        if not reason:
            return
        data = {
            'name': current_user_name,
            'id': current_user_id,
            'reason': reason,
            'manager': employee_info[current_user_id]['manager'],
            'time': datetime.now().strftime("%H:%M:%S"),
            'status': 'submitted',
            'created_by': current_user_id,
            'manager_status': '',
            'hr_status': ''
        }
        save_data(data)
        save_to_excel(reason)
        self.manager.current = 'main'

class ManagerScreen(Screen):
    pending_request = None

    def on_pre_enter(self):
        self.pending_request = None
        requests = load_data()
        for req in requests:
            if req.get('status') == 'submitted':
                if current_user_id == employee_info.get(req['created_by'], {}).get('manager'):
                    self.pending_request = req
                    self.ids.label_review.text = (
                        f"Name: {req['name']}\n"
                        f"ID: {req['id']}\n"
                        f"Reason: {req['reason']}\n"
                        f"Time: {req['time']}"
                    )
                    return
        self.ids.label_review.text = "No pending request."

    def approve(self, status=True):
        if not self.pending_request:
            self.ids.label_review.text = "No pending request."
            return

        all_data = load_data()
        for req in all_data:
            if req == self.pending_request:
                req['status'] = 'manager_approved' if status else 'rejected_by_manager'
                req['manager_status'] = 'Approved' if status else 'Rejected'
                break
        with open(CSV_FILE, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=all_data[0].keys())
            writer.writeheader()
            writer.writerows(all_data)
        self.on_pre_enter()

class HRScreen(Screen):
    pending_request = None

    def on_pre_enter(self):
        self.pending_request = None
        requests = load_data()
        for req in requests:
            if req.get('status') == 'manager_approved' and current_user_id in get_hr_chain(req['created_by']):
                self.pending_request = req
                self.ids.label_hr.text = (
                    f"Name: {req['name']}\n"
                    f"ID: {req['id']}\n"
                    f"Reason: {req['reason']}\n"
                    f"Time: {req['time']}"
                )
                return
        self.ids.label_hr.text = "No HR request."

    def hr_handle(self, approve=True):
        if not self.pending_request:
            self.ids.label_hr.text = "No HR request."
            return
        all_data = load_data()
        for req in all_data:
            if req == self.pending_request:
                if approve:
                    req['status'] = 'hr_approved'
                    req['hr_status'] = f"Approved by {current_user_id}"
                    q = qrcode.make(json.dumps(req))
                    q.save(QR_IMAGE)
                else:
                    req['status'] = 'rejected_by_hr'
                    req['hr_status'] = f"Rejected by {current_user_id}"
                break
        with open(CSV_FILE, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=all_data[0].keys())
            writer.writeheader()
            writer.writerows(all_data)
        self.on_pre_enter()


class GatePassScreen(Screen):
    def on_pre_enter(self):
        requests = load_data()
        for req in requests:
            if req.get('status') == 'hr_approved':
                self.ids.label_pass.text = (
                    f"Name: {req['name']}\nID: {req['id']}\nReason: {req['reason']}\nTime: {req['time']}"
                )
                self.ids.qr_image.source = QR_IMAGE
                return
        self.ids.label_pass.text = 'Gate Pass not ready.'
        self.ids.qr_image.source = ''

    def download_qr(self):
        # For simplicity, just copy the file to a 'downloads' folder
        import shutil

        dest_folder = 'downloads'
        os.makedirs(dest_folder, exist_ok=True)

        dest_file = os.path.join(dest_folder, f'gatepass_qr_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png')

        try:
            shutil.copy(QR_IMAGE, dest_file)
            print(f"QR saved to {dest_file}")
            self.ids.label_pass.text = f"✅ QR saved to {dest_file}"
        except Exception as e:
            print("Download error:", e)
            self.ids.label_pass.text = "❌ Failed to save QR"


class ScanScreen(Screen):
    scanned = False  # To track if already scanned
    current_qr_data = None  # To hold scanned data

    def grant(self):
        if self.scanned:
            self.ids.scan_label.text = "✅ Already scanned. Click OUT to finish."
            return

        qr_data = scan_qr()
        if qr_data:
            try:
                data = json.loads(qr_data)
                self.current_qr_data = data
                self.ids.scan_label.text = (
                    f"✅ SCAN SUCCESS ✅\n\n"
                    f"Name: {data['name']}\n"
                    f"ID: {data['id']}\n"
                    f"Reason: {data['reason']}\n"
                    f"Time: {data['time']}\n"
                    f"Manager Status: {data.get('manager_status', 'N/A')}\n"
                    f"HR Status: {data.get('hr_status', 'N/A')}\n"
                )
                self.scanned = True
            except Exception as e:
                self.ids.scan_label.text = f"❌ Decode error: {e}"
        else:
            self.ids.scan_label.text = "❌ No QR found."

    def checkout(self):
        if not self.scanned or not self.current_qr_data:
            self.ids.scan_label.text = "❌ Scan first before OUT."
            return

        out_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ids.scan_label.text += f"\n\n✅ OUT recorded at {out_time}"

        save_checkout_to_excel(
            self.current_qr_data['id'],
            self.current_qr_data['name'],
            self.current_qr_data['reason']
        )

        self.scanned = False
        self.current_qr_data = None
def save_checkout_to_excel(emp_id, emp_name, reason):
    filename = "gatepass_checkout_log.xlsx"
    headers = ["Date", "Time", "Emp-id", "Emp-name", "Reason", "Status"]

    row = [
        datetime.now().strftime("%Y-%m-%d"),
        datetime.now().strftime("%H:%M:%S"),
        emp_id,
        emp_name,
        reason,
        "OUT"
    ]

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(headers)

    sheet.append(row)
    wb.save(filename)




def scan_qr():
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        for barcode in decode(frame):
            qr_data = barcode.data.decode('utf-8')
            cap.release()
            cv2.destroyAllWindows()
            return qr_data

        cv2.imshow('QR Scanner', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    return None

class GatePassApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(LoginScreen(name='login'))
        sm.add_widget(MainScreen(name='main'))
        sm.add_widget(EntryScreen(name='entry'))
        sm.add_widget(ManagerScreen(name='manager'))
        sm.add_widget(HRScreen(name='hr'))
        sm.add_widget(GatePassScreen(name='gatepass'))
        sm.add_widget(ScanScreen(name='scan'))
        return sm

if __name__ == '__main__':
    GatePassApp().run()
